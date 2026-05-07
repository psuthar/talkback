"""Deterministic e2e fixture generator for spreadsheet upload tests.

Outputs (to the same directory as this script):
- test.csv  : 5 rows (1 header + 4 body) × 3 columns of fixed text
- test.xlsx : 2 sheets, each with 3 rows × 2 columns of fixed text
- test.xls  : same shape as test.xlsx in the legacy binary format

Run with the project venv (or any Python ≥ 3.11 with openpyxl + xlwt installed):

    /tmp/sidecar-test-venv/bin/pip install openpyxl xlwt
    /tmp/sidecar-test-venv/bin/python web/tests/e2e/fixtures/generate_spreadsheet_fixtures.py

Re-running is safe; the binaries are byte-identical because every value is
hard-coded and timestamps are zeroed out where the libraries allow it. The
binaries are checked into git so e2e runs do not require Python.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import xlwt

HERE = Path(__file__).resolve().parent

CSV = """name,role,location
Alex Chen,Senior Staff Engineer,San Francisco
Priya Raman,Hiring Manager,New York
Sara Okafor,Principal Engineer,Seattle
Diego Alvarez,Staff Engineer,Austin
"""

SHEETS = {
    "Roster": [
        ["name", "role"],
        ["Alex Chen", "Senior Staff Engineer"],
        ["Priya Raman", "Hiring Manager"],
    ],
    "Notes": [
        ["topic", "summary"],
        ["culture", "candidate values written communication"],
        ["scope", "leads notification platform"],
    ],
}


def write_csv() -> Path:
    out = HERE / "test.csv"
    out.write_text(CSV)
    return out


def write_xlsx() -> Path:
    wb = openpyxl.Workbook()
    # remove the default sheet that openpyxl creates on init
    wb.remove(wb.active)
    for name, rows in SHEETS.items():
        sheet = wb.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
    out = HERE / "test.xlsx"
    wb.save(out)
    return out


def write_xls() -> Path:
    wb = xlwt.Workbook()
    for name, rows in SHEETS.items():
        sheet = wb.add_sheet(name)
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                sheet.write(r, c, value)
    out = HERE / "test.xls"
    wb.save(str(out))
    return out


def main() -> None:
    for fn in (write_csv, write_xlsx, write_xls):
        path = fn()
        print(f"wrote {path.relative_to(HERE.parent.parent.parent.parent)}  ({path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
